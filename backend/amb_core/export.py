from __future__ import annotations
import html, json, math
from pathlib import Path
from .memo import render_markdown
from .metrics import METRIC_KEYS
from .scoring import DIRECTION, DEFAULT_WEIGHTS, _resolve, _test
from .models import Memo

_PCT={"ann_return","ann_vol","max_drawdown","downside_dev","alpha","tracking_error","hit_rate"}
def _pct(x): return "—" if x is None else f"{x*100:.1f}%"
def _num(x): return "—" if x is None else f"{x:.2f}"

def write_markdown(memo, path):
    p=Path(path);p.parent.mkdir(parents=True,exist_ok=True);p.write_text(render_markdown(memo));return p
def write_json(memo, path):
    p=Path(path);p.parent.mkdir(parents=True,exist_ok=True);p.write_text(memo.model_dump_json(indent=2));return p

_ASSETS = Path(__file__).resolve().parent / "assets"


def _read_asset(name):
    """Load a UI asset (CSS/JS) that lives alongside this module."""
    return (_ASSETS / name).read_text(encoding="utf-8")


def _json_for_script(obj) -> str:
    """Serialize `obj` for safe embedding inside a <script> tag.

    Fund names, strategies and LLM prose flow into this payload, so neutralize
    any sequence that could break out of the script element ("</script>", "<!--")
    or the JSON-in-HTML context. The escaped code points parse back to the exact
    same characters in the browser, so the data the page sees is unchanged.
    """
    return (
        json.dumps(obj, ensure_ascii=False)
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
        .replace("&", "\\u0026")
        .replace("\u2028", "\\u2028")
        .replace("\u2029", "\\u2029")
    )


_CSS = _read_asset("memo.css")

_JS = _read_asset("memo.js")

def _detail_html(sec, full, e):
    cells="".join(f'<div class="cell"><b>{(_pct(full.get(k)) if k in _PCT else _num(full.get(k)))}</b><i>{k.replace("_"," ")}</i></div>' for k in METRIC_KEYS)
    chips=""
    if sec:
        for c in sec.claims:
            k="ok" if c.verified else "warn";src=e("; ".join(c.source_refs) or "no source")
            chips+=f'<span class="{k}" title="{src}"> {e(c.text)}</span>'
    body=e(sec.body) if sec else ""
    return (f'<p class="d-p">{body}</p><div class="mgrid">{cells}</div>'
            f'<div class="src-lbl">Sources · verified against the metrics engine</div><div class="chipx">{chips}</div>')

def _reason(f, mandate):
    strat=f.strategy; vol=None
    for c in mandate.constraints:
        if c.field=="strategy" and c.op=="not_in" and strat in (c.value or []):
            return f"illiquid · {strat}", "LIQUIDITY"
    return None, None

def render_html(memo, ctx=None):
    e=html.escape;a=memo.audit;sl=memo.shortlist
    ranks={s.fund_id:s.rank for s in sl}
    series=(ctx.series_by_fund if ctx else {}) or {}
    mbf=(ctx.metrics_by_fund if ctx else {s.fund_id:s.metrics for s in sl})
    mandate=ctx.mandate if ctx else None
    secs={}
    for i,s in enumerate(sl): secs[s.fund_id]=memo.sections[1+i] if 1+i<len(memo.sections) else None

    # score components — z across the ELIGIBLE set, the SAME basis build_shortlist
    # ranks on, so the displayed bar score a fund shows is the score it's ranked on.
    weights=(mandate.weights if (mandate and mandate.weights) else DEFAULT_WEIGHTS)
    import statistics
    from .scoring import apply_constraints
    _usable=[f for f in ctx.funds.values() if f.fund_id in mbf] if ctx else []
    _elig=(set(apply_constraints(_usable, mbf, mandate)) if (ctx and mandate) else set(ranks)) or set(ranks)
    stats={}
    for k in weights:
        vals=[mbf[fid].get(k) for fid in _elig if mbf.get(fid,{}).get(k) is not None]
        if len(vals)>=2: stats[k]=(statistics.fmean(vals),statistics.pstdev(vals))
    def components(fid):
        out=[]
        for k,w in weights.items():
            v=mbf[fid].get(k)
            if v is None or k not in stats: continue
            mu,sd=stats[k]
            if sd==0: continue
            out.append({"k":k,"c":round(w*((v-mu)/sd)*DIRECTION.get(k,0),3)})
        return out

    # benchmark point (same conventions as the metrics engine: CAGR + sample-std annualized vol)
    bench=None
    bmk=getattr(ctx,"benchmark",None) if ctx else None
    if bmk and getattr(bmk,"points",None):
        bvals=bmk.values; nb=len(bvals); bppy=bmk.periods_per_year or 12
        if nb>=2:
            growth=1.0
            for v in bvals: growth*=(1.0+v)
            bret=(growth**(bppy/nb)-1.0) if growth>0 else growth-1.0
            bmean=sum(bvals)/nb
            bvar=sum((v-bmean)**2 for v in bvals)/(nb-1)
            bvol=(bvar**0.5)*(bppy**0.5)
            bwealth=[];bc=1.0
            for v in bvals: bc*=(1.0+v);bwealth.append(round(bc,4))
            bench={"name":bmk.name,"vol":bvol,"ret":bret,"wealth":bwealth,
                   "kind":getattr(bmk,"source_kind","snapshot"),"srcName":getattr(bmk,"source_name","snapshot"),
                   "asOf":str(getattr(bmk,"as_of","")),"n":nb}

    # gates — one per hard constraint, with an allocator-legible label
    def _gate_meta(c):
        if c.field=="strategy" and c.op=="not_in":
            vals=", ".join(map(str,c.value or [])) if isinstance(c.value,(list,tuple)) else str(c.value)
            return {"label":"STRATEGY","detail":f"excl. {vals}","kind":"STRATEGY"}
        if c.field=="redemption_days" and c.op in ("<=","<"):
            return {"label":"LIQUIDITY","detail":f"≤ {int(c.value)}d","kind":"LIQUIDITY"}
        if c.field=="ann_vol" and c.op in ("<=","<"):
            return {"label":"VOLATILITY","detail":f"≤ {c.value*100:.0f}%","kind":"VOLATILITY"}
        if c.field=="max_drawdown" and c.op in (">=",">"):
            return {"label":"DRAWDOWN","detail":f"≥ {c.value*100:.0f}%","kind":"DRAWDOWN"}
        return {"label":c.field.upper()[:9],"detail":f"{c.op} {c.value}","kind":c.field.upper()[:9]}
    gates=[{"label":g["label"],"detail":g["detail"]} for g in (map(_gate_meta,mandate.constraints) if mandate else [])]
    if not gates: gates=[{"label":"MANDATE","detail":"screen"}]

    def _one_reason(f,m,c):
        g=_gate_meta(c)
        val=_resolve(f,m,c.field)
        if c.field=="strategy":
            return {"text":f"off-strategy · {f.strategy}","kind":g["kind"]}
        if c.field=="redemption_days":
            fr=(f.redemption_freq or "illiquid")
            dd=(f"{int(val)}d" if val is not None else "n/a")
            return {"text":f"illiquid · {fr} ({dd})","kind":g["kind"]}
        if c.field=="ann_vol":
            return {"text":f"too volatile · {m.get('ann_vol',0)*100:.0f}% > {c.value*100:.0f}% cap","kind":g["kind"]}
        if c.field=="max_drawdown":
            return {"text":f"drawdown · {m.get('max_drawdown',0)*100:.0f}% beyond {c.value*100:.0f}% floor","kind":g["kind"]}
        return {"text":f"fails {c.field} {c.op} {c.value}","kind":g["kind"]}

    def _reject_reasons(f,m):
        """EVERY failing hard constraint -> list of {text,kind}. [] if eligible.
        A fund can breach several limits (a venture sleeve is illiquid AND too
        drawdown-prone AND off-strategy); the funnel shows each so every gate does
        visible work."""
        if not mandate or f is None:
            return []
        out=[]
        for c in mandate.constraints:
            val=_resolve(f,m,c.field)
            if not _test(val,c.op,c.value):
                out.append(_one_reason(f,m,c))
        return out

    volcap=None
    if mandate:
        volcap=next((c.value for c in mandate.constraints if c.field=="ann_vol" and c.op=="<="),None)
    gateX=None
    plist=[(fid,m) for fid,m in mbf.items() if m.get("ann_vol") is not None and m.get("ann_return") is not None]
    fd=[]
    if plist:
        vols=[m["ann_vol"] for _,m in plist];rets=[m["ann_return"] for _,m in plist]
        vmin,vmax=min(vols),max(vols);rmin,rmax=min(rets),max(rets);vr=(vmax-vmin) or 1;rr=(rmax-rmin) or 1
        for fid,m in plist:
            f=ctx.get_fund(fid) if ctx else None;ser=series.get(fid)
            wealth=[];c=1.0
            if ser:
                for v in ser.values:c*=(1+v);wealth.append(round(c,4))
            rk=ranks.get(fid)
            reasons=_reject_reasons(f,m)
            reason=(reasons[0]["text"] if reasons else None)
            reason_kind=(reasons[0]["kind"] if reasons else None)
            eligible=(reason is None)                 # passed the mandate (ranked OR outscored)
            cut=(rk is None and eligible)             # eligible but below the top-N shortlist
            comps=components(fid) if eligible else []
            netret=(ctx.net_return(fid) if ctx else None)
            fd.append({"id":fid,"name":(f.name if f else fid),"strategy":(f.strategy if f else ""),
                "rank":rk,"excluded":rk is None,"eligible":eligible,"cut":cut,"rkind":reason_kind,
                "srank":(rk if rk else (90 if cut else 99)),
                "x":round(12+(m["ann_vol"]-vmin)/vr*76,1),"y":round(12+(m["ann_return"]-rmin)/rr*76,1),
                "ret":m.get("ann_return"),"vol":m.get("ann_vol"),"sharpe":m.get("sharpe"),"sortino":m.get("sortino"),"calmar":m.get("calmar"),"maxdd":m.get("max_drawdown"),
                "beta":m.get("beta"),"alpha":m.get("alpha"),"corr":m.get("correlation"),
                "fee":(f.mgmt_fee_pct if f else None),"netret":netret,
                "redf":(f.redemption_freq if f else None),"redd":(f.redemption_days if f else None),
                "lockup":(f.lockup_months if f else None),"notice":(f.notice_days if f else None),
                "wealth":wealth,"reason":reason,"reasons":reasons,"components":comps,"comp":{x["k"]:x["c"] for x in comps},"score":round(sum(x["c"] for x in comps),3),
                "detail":_detail_html(secs.get(fid),mbf.get(fid,{}),e)})
    # zoom positions (all mandate-eligible funds + benchmark, shared range so the
    # frontier reads against the index and the corners carry meaning)
    surv=[d for d in fd if d["eligible"]]
    benchLine=None
    if surv:
        zv=[d["vol"] for d in surv];zr=[d["ret"] for d in surv]
        if bench: zv=zv+[bench["vol"]];zr=zr+[bench["ret"]]
        zvmin,zvmax=min(zv),max(zv);zrmin,zrmax=min(zr),max(zr);zvr=(zvmax-zvmin) or 1;zrr=(zrmax-zrmin) or 1
        for d in surv: d["xz"]=round(14+(d["vol"]-zvmin)/zvr*72,1);d["yz"]=round(14+(d["ret"]-zrmin)/zrr*72,1)
        if bench:
            bench["xz"]=round(14+(bench["vol"]-zvmin)/zvr*72,1)
            bench["yz"]=round(14+(bench["ret"]-zrmin)/zrr*72,1)
            if bench["vol"]>0:
                s=bench["ret"]/bench["vol"]  # index return-per-unit-risk
                def _mapxy(vol):
                    return (round(14+(vol-zvmin)/zvr*72,1), round(14+(s*vol-zrmin)/zrr*72,1))
                x1,y1=_mapxy(zvmin);x2,y2=_mapxy(zvmax)
                benchLine={"x1":x1,"y1":y1,"x2":x2,"y2":y2}
    for d in fd: d.setdefault("xz",d["x"]);d.setdefault("yz",d["y"])

    if plist and volcap is not None:
        gx=12+(volcap-vmin)/vr*76
        if 0<gx<100: gateX=round(gx,1)
    top=sl[0] if sl else None
    vplain=f"{top.name} leads on risk-adjusted return." if top else "No fund met the mandate."
    vhtml=(f"<b>{e(top.name)}</b> leads on risk-adjusted return." if top else "No fund met the mandate.")
    def _fname(fid):
        f=ctx.get_fund(fid) if ctx else None; return f.name if f else fid
    audit_claims=[{"fund":_fname(c.get("fund_id")),"metric":(c.get("metric") or "").replace("_"," "),
                   "value":c.get("value"),"verified":bool(c.get("verified")),"sources":c.get("sources",[])}
                  for c in a.get("claims",[])]
    share_txt=((top.name+" — EQUI Investment Committee recommendation (leads on risk-adjusted return). Shortlist: "
                +"; ".join(f"{s.rank}. {s.name} ({_pct(s.metrics.get('ann_return'))})" for s in sl)
                +f". {a.get('verified_count',0)}/{a.get('claim_count',0)} claims verified against the metrics engine.") if top else "No fund met the mandate.")
    # memo prose sections (Summary / Recommendation / Key Risks / Data Appendix)
    def _find_sec(h):
        return next((s for s in memo.sections if s.heading==h), None)
    _smy=_find_sec("Summary");_rec=_find_sec("Recommendation");_kr=_find_sec("Key Risks");_apx=_find_sec("Data Appendix")
    kr_claims=[{"fund":_fname(c.fund_id),"metric":(c.metric or "").replace("_"," "),"value":c.value,
                "verified":bool(c.verified),"text":e(c.text)} for c in (_kr.claims if _kr else [])]
    memo_payload={"summary":e(_smy.body) if _smy else "","recommendation":e(_rec.body) if _rec else "",
                  "keyRisks":{"body":e(_kr.body) if _kr else "","claims":kr_claims},
                  "appendix":e(_apx.body) if _apx else ""}
    rd=getattr(ctx,"readiness",{}) if ctx else {}
    liqCap=(next((c.value for c in mandate.constraints if c.field=="redemption_days" and c.op in ("<=","<")),None) if mandate else None)
    maxddFloor=(next((c.value for c in mandate.constraints if c.field=="max_drawdown" and c.op in (">=",">")),None) if mandate else None)
    DATA={"funds":fd,"gates":gates,"verdict":vplain,"verdictHtml":vhtml,"model":memo.generated_by,"title":memo.title,
          "verified":a.get("verified_count",0),"total":a.get("claim_count",0),"mandate":memo.mandate,"gateX":gateX,"volcap":(f"{volcap*100:.0f}%" if volcap is not None else None),"weights":{k:round(float(v),3) for k,v in weights.items()},
          "bench":({k:(round(v,4) if isinstance(v,float) else v) for k,v in bench.items()} if bench else None),"benchLine":benchLine,
          "nTotal":len(fd),"nEligible":sum(1 for d in fd if d["eligible"]),"nShort":len(sl),"nReject":sum(1 for d in fd if d.get("reason")),
          "audit":audit_claims,"shareText":share_txt,"readiness":rd,"memo":memo_payload,
          "rfUsed":(getattr(ctx,"rf_used",None) if ctx else None),"rfSource":(getattr(ctx,"rf_source","mandate") if ctx else "mandate"),
          "mandateSpec":{"exclStrats":(next((c.value for c in mandate.constraints if c.field=="strategy" and c.op=="not_in"),[]) if mandate else []) or [],
                         "volCap":(next((c.value for c in mandate.constraints if c.field=="ann_vol" and c.op=="<="),None) if mandate else None),
                         "liqCap":liqCap,"maxddFloor":maxddFloor,
                         "rf":(mandate.risk_free_annual if mandate else 0.02),"topN":(mandate.top_n if mandate else 5)},
          "dir":{k:DIRECTION.get(k,0) for k in weights}}

    chips="".join(f'<div class="chip{" r1" if s.rank==1 else ""}" data-fid="{e(s.fund_id)}" title="Open fund detail"><span class="n">{s.rank:02d}</span><span class="nm">{e(s.name)}</span><span class="rt">{_pct(s.metrics.get("ann_return"))}</span><span class="cx">⤢</span></div>' for s in sl)

    # ── designed print / PDF document (screen-hidden; shown only in @media print) ──
    win=sl[0] if sl else None
    def _mkpi(lb,vl): return f'<div class="pd-kpi"><b>{vl}</b><i>{lb}</i></div>'
    kpis=""
    if win:
        wm=win.metrics
        kpis=(_mkpi("Ann. return",_pct(wm.get("ann_return")))+_mkpi("Volatility",_pct(wm.get("ann_vol")))
             +_mkpi("Sharpe",_num(wm.get("sharpe")))+_mkpi("Sortino",_num(wm.get("sortino")))
             +_mkpi("Calmar",_num(wm.get("calmar")))+_mkpi("Max drawdown",_pct(wm.get("max_drawdown"))))
    trows=""
    for s in sl:
        m=s.metrics; wc=' class="win"' if s.rank==1 else ''
        trows+=(f'<tr{wc}><td class="r">{s.rank:02d}</td><td class="nm">{e(s.name)}</td><td class="st">{e(s.strategy)}</td>'
                f'<td>{_pct(m.get("ann_return"))}</td><td>{_pct(m.get("ann_vol"))}</td><td>{_num(m.get("sharpe"))}</td>'
                f'<td>{_num(m.get("sortino"))}</td><td>{_num(m.get("calmar"))}</td><td>{_pct(m.get("max_drawdown"))}</td>'
                f'<td class="sc">{s.score:+.2f}</td></tr>')
    exrows="".join(f'<tr><td class="nm">{e(d["name"])}</td><td class="st">{e(d["strategy"])}</td><td class="ex">{e(d["reason"])}</td><td>{_pct(d["ret"])}</td><td>{_pct(d["vol"])}</td></tr>' for d in fd if d.get("reason"))
    screens=" &nbsp;·&nbsp; ".join(f'{g["label"].title()} {g["detail"]}' for g in gates)
    wtext=" &nbsp;·&nbsp; ".join(f'{k.replace("_"," ")} {round(v*100)}%' for k,v in weights.items())
    benchtext=(f'{e(bench["name"])} — ann. return {_pct(bench["ret"])} · volatility {_pct(bench["vol"])}' if bench else "—")
    printdoc=(f'<div id="printdoc"><div class="pd-top"><div class="pd-brand"><div class="pd-logo"></div><span>EQUI</span></div>'
              f'<div class="pd-meta">Investment Committee · Recommendation<br><b id="pd-date"></b></div></div>'
              f'<div class="pd-verdict">{vhtml}</div>'
              f'<div class="pd-hero"><div class="pd-rec"><div class="pd-lbl">Recommendation</div>'
              f'<div class="pd-name">{e(win.name) if win else "—"}</div><div class="pd-strat">{e(win.strategy) if win else ""}</div></div>'
              f'<div class="pd-kpis">{kpis}</div></div>'
              f'<div class="pd-two"><div class="pd-card"><div class="pd-sec">Mandate · hard limits</div><div class="pd-body">{screens}</div>'
              f'<div class="pd-sec2">Scoring weights</div><div class="pd-body">{wtext}</div></div>'
              f'<div class="pd-card"><div class="pd-sec">Measured against</div><div class="pd-body">{benchtext}</div>'
              f'<div class="pd-sec2">Method</div><div class="pd-body">Funds clearing both limits are scored on six risk-adjusted measures; the top {len(sl)} advance and the leader is recommended.</div></div></div>'
              f'<div class="pd-sec">Shortlist · ranked by weighted risk-adjusted score</div>'
              f'<table class="pd-tbl"><thead><tr><th class="r">#</th><th class="nm">Fund</th><th class="st">Strategy</th><th>Return</th><th>Vol</th><th>Sharpe</th><th>Sortino</th><th>Calmar</th><th>Max DD</th><th class="sc">Score</th></tr></thead><tbody>{trows}</tbody></table>'
              + (f'<div class="pd-sec">Excluded by mandate</div><table class="pd-tbl pd-ex"><tbody>{exrows}</tbody></table>' if exrows else '')
              + f'<div class="pd-foot"><span class="pd-ck">✓</span> Every figure re-verified against the deterministic metrics engine — {a.get("verified_count",0)}/{a.get("claim_count",0)} claims verified · {"offline deterministic build" if memo.generated_by=="template" else "generated by "+e(memo.generated_by)}</div></div>')

    header=(f'<div class="hdr"><div class="brand"><div class="logo"></div><div class="wm">EQUI</div>'
            '<div class="theme" id="themebtn" role="button" tabindex="0" aria-label="Toggle light or dark theme" aria-pressed="false"><i class="tg tg-d">☾</i><i class="tg tg-l">☀︎</i><span class="knob"></span></div></div>'
            f'<div class="vwrap"><div class="vpre">Investment Committee · Recommendation</div><div class="vtext" id="vtext"></div></div>'
            f'<div class="right">'
            f'<button class="hbtn hbtn-pause" id="pausebtn" title="Pause or resume the replay (spacebar)">❚❚&nbsp;pause</button>'
            f'<button class="hbtn" id="skip">skip replay ▸</button>'
            f'<button class="hbtn" id="liveBtn" title="Fetch live market data from FRED (requires ./amb serve)">↻ live data</button>'
            f'<button class="hbtn" id="mandateBtn" title="Adjust the mandate — constraints and scoring weights — and re-decide live">mandate</button>'
            f'<button class="hbtn" id="memoBtn" title="Read the full IC memo — summary, recommendation, key risks, data appendix">memo</button>'
            f'<button class="hbtn" id="upBtn" title="Load a fund-universe CSV (funds + returns) and re-run the analysis in place">load csv</button>'
            f'<input type="file" id="upInput" accept=".csv" multiple style="display:none">'
            f'<button class="hbtn" id="shareBtn" title="Share the recommendation summary">share</button>'
            f'<button class="hbtn hbtn-primary" id="dlBtn" title="Download the memo as a PDF">download</button>'
            f'<button class="shieldbtn" id="vbadge" role="button" tabindex="0" title="{a.get("verified_count",0)}/{a.get("claim_count",0)} claims verified against the metrics engine — click for the audit trail">'
            f'<svg viewBox="0 0 24 24" fill="none"><path d="M12 2.4l7 2.9v5.7c0 4.7-3.3 8-7 9.6-3.7-1.6-7-4.9-7-9.6V5.3l7-2.9z" fill="var(--accent-soft)" stroke="var(--accent2)" stroke-width="1.4" stroke-linejoin="round"/><path d="M8.6 12.2l2.3 2.3 4.5-4.6" stroke="var(--accent2)" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"/></svg>'
            f'<i class="vdot"></i></button>'
            f'</div></div>')

    stage=('<div class="stage"><div id="field"><div id="fieldwash"></div><div class="sweetz"><span>sweet spot · high return / low risk</span></div>'
           '<div id="guides"><div id="benchray"></div><div id="benchmk"><div class="dm"></div><div class="bl"></div></div><div id="beatlbl"></div></div>'
           '<div class="gates" id="gates"></div><div class="gateline" id="gateline"><span></span></div><div class="danger" id="danger"></div><div id="counter"></div><div id="factorcue"><span class="fdot"></span><span class="fname"></span><span class="fwt"><i></i></span><span class="fpct"></span></div><div class="ax y">Return →</div><div class="ax x">Risk · volatility →</div></div>'
           '<div id="chapter"></div><div id="stagepaused">❚❚ paused</div><div id="srcchip" title=""></div></div>')

    side=('<div class="side">'
          '<div id="intropane"><div class="ip-h">The mandate</div><div class="ip-s">what advances · and how it\'s scored</div>'
          '<div class="ip-lbl">Screen · hard limits</div><div id="ip-gates"></div>'
          '<div class="ip-lbl">Score · weighting</div><div id="ip-weights"></div>'
          '<div class="ip-lbl">Measured against</div><div id="ip-bench"></div>'
          f'<div class="ip-note">Funds clearing both limits are scored on the six risk-adjusted measures above. The top {len(sl)} by weighted score advance to the shortlist; the leader is the recommendation.</div>'
          '<div id="ip-tally"></div></div>'
          '<div class="pane" id="trajpane"><div class="head"><div class="t">36-Month Trajectory</div><div class="s">growth of $1 · vs S&amp;P 500</div><span class="srcbadge" id="benchsrc"></span></div><div id="trajwrap"><svg id="traj"></svg></div></div>'
          '<div class="pane" id="scorepane"><div class="head"><div class="t">The weighing</div><div class="s">how the ranking forms</div></div><div id="weighticker"></div><div id="weighlegend"></div><div id="scorebars"></div><div class="why-note" id="whynote"></div></div>'
          '</div>')

    rail=f'<div class="rail"><div class="rl">Shortlist</div><div class="chips">{chips}</div></div>'

    return ('<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">'
            f'<title>{e(memo.title)}</title>'
            '<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>'
            '<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">'
            f'<style>{_CSS}</style></head><body>'
            '<div class="atmo"><div class="grid"></div></div><div id="tip"></div>'
            f'<div class="app">{header}<div class="mid">{stage}{side}</div>{rail}</div>'
            f'{printdoc}'
            '<div id="drawer"></div><div id="play">Replay decision</div><div id="pop"></div><div id="toast"></div>'
            f'<script>window.AMB={_json_for_script(DATA)};</script><script>{_JS}</script></body></html>')

def write_html(memo, path, ctx=None):
    p=Path(path);p.parent.mkdir(parents=True,exist_ok=True);p.write_text(render_html(memo,ctx));return p

def write_xlsx(memo, ctx, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    p=Path(path);p.parent.mkdir(parents=True,exist_ok=True);wb=Workbook()
    hf=PatternFill("solid",fgColor="141821");ff=Font(bold=True,color="84AEC8")
    def sh(ws,n):
        for c in range(1,n+1):ws.cell(row=1,column=c).fill=hf;ws.cell(row=1,column=c).font=ff
    ws=wb.active;ws.title="Shortlist"
    cols=["rank","fund_id","name","strategy","score","ann_return","sharpe","sortino","calmar","max_drawdown","ann_vol"]
    ws.append([c.replace('_',' ').title() for c in cols])
    for s in memo.shortlist:
        m=s.metrics;ws.append([s.rank,s.fund_id,s.name,s.strategy,s.score,m.get('ann_return'),m.get('sharpe'),m.get('sortino'),m.get('calmar'),m.get('max_drawdown'),m.get('ann_vol')])
    sh(ws,len(cols))
    ws2=wb.create_sheet("Metrics");hdr=["fund_id","name"]+METRIC_KEYS
    ws2.append([h.replace('_',' ').title() for h in hdr])
    for fid,vals in ctx.metrics_by_fund.items():
        f=ctx.get_fund(fid);ws2.append([fid,f.name if f else fid]+[vals.get(k) for k in METRIC_KEYS])
    sh(ws2,len(hdr))
    ws3=wb.create_sheet("Audit");ws3.append(["fund_id","metric","value","verified","sources"])
    for c in memo.audit.get("claims",[]):ws3.append([c.get("fund_id"),c.get("metric"),c.get("value"),c.get("verified"),"; ".join(c.get("sources",[]))])
    sh(ws3,5);wb.save(p);return p

def export_all(memo, ctx=None, outdir="exports"):
    paths={"md":write_markdown(memo,Path(outdir)/"memo.md"),"json":write_json(memo,Path(outdir)/"memo.json"),"html":write_html(memo,Path(outdir)/"memo.html",ctx)}
    if ctx is not None:paths["xlsx"]=write_xlsx(memo,ctx,Path(outdir)/"memo.xlsx")
    return paths
