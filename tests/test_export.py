from openpyxl import load_workbook
from amb_core.export import export_all, render_html


def test_all_formats_written(sample_run, tmp_path):
    memo, ctx = sample_run
    paths = export_all(memo, ctx, tmp_path)
    for k in ["md", "json", "html", "xlsx"]:
        assert paths[k].exists() and paths[k].stat().st_size > 0


def test_html_has_map_and_verified(sample_run):
    memo, ctx = sample_run
    h = render_html(memo, ctx)
    assert memo.shortlist[0].name in h
    assert "verified" in h
    assert 'id="field"' in h and "window.AMB" in h  # interactive field + data payload


def test_xlsx_has_expected_sheets(sample_run, tmp_path):
    memo, ctx = sample_run
    wb = load_workbook(export_all(memo, ctx, tmp_path)["xlsx"])
    assert {"Shortlist", "Metrics", "Audit"}.issubset(set(wb.sheetnames))
